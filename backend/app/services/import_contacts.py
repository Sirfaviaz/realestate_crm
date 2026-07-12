import uuid
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.contact import Contact
from app.models.listing import Listing
from app.services.matching import match_all_active

CONTACT_FIELDS = ["name", "phone", "email", "whatsapp"]
CRM_FIELDS = ["urgency", "lead_score", "notes"]

PROJECT_FIELDS = [
    "builder_name",
    "project_name",
    "location",
    "project_type",
    "land_area_cent",
    "launching_time",
    "completion_time",
    "total_floors",
    "parking_details",
    "bhk",
    "sqft",
    "available_floors",
    "amenities",
    "base_price",
    "car_parking_price",
    "gst_percent",
    "down_payment_percent",
    "utility_charge",
    "total_instalment",
    "total_amount",
    "price_as_of_date",
]

SELLER_SALE_FIELDS = PROJECT_FIELDS

LANDLORD_FIELDS = PROJECT_FIELDS + ["monthly_rent", "security_deposit"]

TEMPLATES: dict[str, dict] = {
    "buyer": {
        "required": ["name", "phone"],
        "optional": [
            "email",
            "whatsapp",
            "budget_min",
            "budget_max",
            "preferred_bhk",
            "preferred_locations",
            "urgency",
            "lead_score",
            "notes",
        ],
        "stream": "sales",
        "role": "buyer",
        "creates_listing": False,
    },
    "seller": {
        "required": ["name", "phone", "project_name"],
        "optional": ["email", "whatsapp", *SELLER_SALE_FIELDS, *CRM_FIELDS],
        "stream": "sales",
        "role": "seller",
        "creates_listing": True,
    },
    "renter": {
        "required": ["name", "phone"],
        "optional": [
            "email",
            "whatsapp",
            "rent_budget",
            "preferred_locations",
            "preferred_bhk",
            "move_in_date",
            "tenant_type",
            "occupant_count",
            "profession",
            "workplace_text",
            "urgency",
            "lead_score",
            "notes",
        ],
        "stream": "rental",
        "role": "renter",
        "creates_listing": False,
    },
    "landlord": {
        "required": ["name", "phone", "project_name"],
        "optional": ["email", "whatsapp", *LANDLORD_FIELDS, *CRM_FIELDS],
        "stream": "rental",
        "role": "landlord",
        "creates_listing": True,
    },
}

HEADER_ALIASES: dict[str, str] = {
    "s no": "sno",
    "sno": "sno",
    "builder": "builder_name",
    "project name": "project_name",
    "location": "location",
    "project type": "project_type",
    "land area (cent)": "land_area_cent",
    "land area cent": "land_area_cent",
    "launching time": "launching_time",
    "completion time": "completion_time",
    "total floors": "total_floors",
    "parking details": "parking_details",
    "type of apartment (bhks)": "bhk",
    "type of apartment bhks": "bhk",
    "bhk": "bhk",
    "sqft of each apartments": "sqft",
    "sqft": "sqft",
    "floors of apartment (available)": "available_floors",
    "floors of apartment available": "available_floors",
    "common amenities": "amenities",
    "basic price floor price": "base_price",
    "basic price": "base_price",
    "car parking price": "car_parking_price",
    "gst (%)": "gst_percent",
    "gst percent": "gst_percent",
    "gst": "gst_percent",
    "down payment (%)": "down_payment_percent",
    "down payment percent": "down_payment_percent",
    "utility charge": "utility_charge",
    "total instalment": "total_instalment",
    "total installment": "total_instalment",
    "total amount": "total_amount",
    "as per date": "price_as_of_date",
    "price as of date": "price_as_of_date",
    "property location": "location",
    "property_location": "location",
    "asking price": "total_amount",
    "asking_price": "total_amount",
    "preferred bhk": "preferred_bhk",
    "preferred_bhk": "preferred_bhk",
    "preferred locations": "preferred_locations",
    "preferred_locations": "preferred_locations",
    "budget min": "budget_min",
    "budget max": "budget_max",
    "rent budget": "rent_budget",
    "move in date": "move_in_date",
    "monthly rent": "monthly_rent",
    "security deposit": "security_deposit",
    "lead score": "lead_score",
}

FLOAT_FIELDS = {
    "budget_min",
    "budget_max",
    "rent_budget",
    "asking_price",
    "sqft",
    "land_area_cent",
    "base_price",
    "car_parking_price",
    "gst_percent",
    "down_payment_percent",
    "utility_charge",
    "total_instalment",
    "total_amount",
    "monthly_rent",
    "security_deposit",
}
INT_FIELDS = {"total_floors", "occupant_count"}
DATE_FIELDS = {"move_in_date", "timeline_date", "price_as_of_date"}
LIST_FIELDS = {"preferred_locations", "amenities"}

_preview_store: dict[str, dict] = {}


def _normalize_header(raw: Any) -> str:
    if raw is None:
        return ""
    key = str(raw).strip().lower().replace("_", " ")
    key = " ".join(key.split())
    key = key.replace(".", "")
    if key in HEADER_ALIASES:
        return HEADER_ALIASES[key]
    return key.replace(" ", "_").replace("(", "").replace(")", "").replace("%", "percent")


def generate_template_workbook(role: str) -> Workbook:
    cfg = TEMPLATES[role]
    wb = Workbook()
    ws = wb.active
    ws.title = role.capitalize()
    headers = cfg["required"] + cfg["optional"]
    ws.append(headers)
    return wb


def _parse_value(key: str, val: Any) -> Any:
    if val is None or val == "":
        return None
    if key in FLOAT_FIELDS:
        try:
            return float(val)
        except (TypeError, ValueError):
            return None
    if key in INT_FIELDS:
        try:
            return int(float(val))
        except (TypeError, ValueError):
            return None
    if key in DATE_FIELDS:
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        try:
            return datetime.fromisoformat(str(val).split("T")[0]).date()
        except (TypeError, ValueError):
            return None
    if key in LIST_FIELDS and isinstance(val, str):
        return [x.strip() for x in val.split(",") if x.strip()]
    if key == "preferred_bhk" and not val:
        return None
    return str(val).strip() if isinstance(val, (str, int, float)) else val


def _listing_fields(data: dict, cfg: dict) -> dict:
    location = data.get("location") or data.get("property_location")
    price = data.get("total_amount") or data.get("base_price") or data.get("asking_price")
    if cfg["role"] == "landlord":
        price = data.get("monthly_rent") or data.get("rent_budget") or price
    title = data.get("project_name") or location or f"{data['name']} property"
    return {
        "stream_type": cfg["stream"],
        "title": title,
        "location_text": location,
        "bhk": data.get("bhk") or data.get("preferred_bhk"),
        "property_type": data.get("project_type"),
        "sqft": data.get("sqft"),
        "price": price,
        "builder_name": data.get("builder_name"),
        "project_name": data.get("project_name"),
        "land_area_cent": data.get("land_area_cent"),
        "launching_time": data.get("launching_time"),
        "completion_time": data.get("completion_time"),
        "total_floors": data.get("total_floors"),
        "parking_details": data.get("parking_details"),
        "available_floors": data.get("available_floors"),
        "amenities": data.get("amenities"),
        "base_price": data.get("base_price"),
        "car_parking_price": data.get("car_parking_price"),
        "gst_percent": data.get("gst_percent"),
        "down_payment_percent": data.get("down_payment_percent"),
        "utility_charge": data.get("utility_charge"),
        "total_instalment": data.get("total_instalment"),
        "total_amount": data.get("total_amount"),
        "price_as_of_date": data.get("price_as_of_date"),
        "monthly_rent": data.get("monthly_rent"),
        "security_deposit": data.get("security_deposit"),
        "status": "available",
        "description": data.get("notes"),
    }


async def preview_import(db: AsyncSession, role: str, file_bytes: bytes) -> dict:
    cfg = TEMPLATES.get(role)
    if not cfg:
        raise ValueError("Invalid role")

    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"preview_id": None, "valid": [], "errors": ["Empty file"], "duplicates": []}

    header_row_idx = 0
    for i, row in enumerate(rows[:10]):
        normalized = [_normalize_header(h) for h in row if h is not None]
        if "name" in normalized or "project_name" in normalized or "phone" in normalized:
            header_row_idx = i
            break

    headers = [_normalize_header(h) for h in rows[header_row_idx]]
    while headers and not headers[-1]:
        headers.pop()

    data_start = header_row_idx + 1
    if data_start < len(rows):
        maybe_keys = [_normalize_header(h) for h in rows[data_start] if h is not None]
        key_hits = sum(1 for k in maybe_keys if k in headers)
        if maybe_keys and key_hits >= max(3, len(headers) // 2):
            data_start += 1

    missing = [r for r in cfg["required"] if r not in headers]
    if missing:
        return {"preview_id": None, "valid": [], "errors": [f"Missing columns: {', '.join(missing)}"], "duplicates": []}

    valid: list[dict] = []
    errors: list[dict] = []
    duplicates: list[dict] = []

    for i, row in enumerate(rows[data_start:], start=data_start + 1):
        data = {headers[j]: row[j] if j < len(row) else None for j in range(len(headers))}
        if not any(v not in (None, "") for v in data.values()):
            continue
        parsed: dict = {}
        row_errors = []
        for key in cfg["required"] + cfg["optional"]:
            if key not in headers:
                continue
            parsed[key] = _parse_value(key, data.get(key))
        if not parsed.get("name"):
            row_errors.append("name required")
        if not parsed.get("phone"):
            row_errors.append("phone required")
        if cfg.get("creates_listing") and not parsed.get("project_name"):
            row_errors.append("project_name required")
        if row_errors:
            errors.append({"row": i, "errors": row_errors, "data": parsed})
            continue

        existing = await db.execute(select(Contact).where(Contact.phone == str(parsed["phone"])))
        dup = existing.scalar_one_or_none()
        entry = {
            "row": i,
            "data": parsed,
            "existing_id": str(dup.id) if dup else None,
            "listing_preview": _listing_fields(parsed, cfg) if cfg.get("creates_listing") else None,
        }
        if dup:
            duplicates.append(entry)
        valid.append(entry)

    preview_id = str(uuid.uuid4())
    _preview_store[preview_id] = {"role": role, "entries": valid, "cfg": cfg}
    return {
        "preview_id": preview_id,
        "valid_count": len(valid),
        "error_count": len(errors),
        "duplicate_count": len(duplicates),
        "valid": valid[:50],
        "errors": errors[:50],
        "duplicates": duplicates[:50],
    }


async def confirm_import(
    db: AsyncSession,
    preview_id: str,
    update_duplicates: bool,
    assigned_user_id: uuid.UUID | None,
) -> dict:
    batch = _preview_store.pop(preview_id, None)
    if not batch:
        raise ValueError("Preview expired or not found")

    cfg = batch["cfg"]
    created = updated = skipped = listings_created = 0

    for entry in batch["entries"]:
        data = entry["data"]
        existing_id = entry.get("existing_id")
        fields = {
            "name": data["name"],
            "phone": str(data["phone"]),
            "email": data.get("email"),
            "whatsapp": data.get("whatsapp"),
            "roles": [cfg["role"]],
            "stream_type": cfg["stream"],
            "budget_min": data.get("budget_min"),
            "budget_max": data.get("budget_max"),
            "rent_budget": data.get("rent_budget") or data.get("monthly_rent"),
            "preferred_bhk": data.get("preferred_bhk") or data.get("bhk"),
            "preferred_locations": data.get("preferred_locations"),
            "property_location": data.get("location") or data.get("property_location"),
            "asking_price": data.get("total_amount") or data.get("base_price") or data.get("asking_price"),
            "sqft": data.get("sqft"),
            "move_in_date": data.get("move_in_date"),
            "timeline_date": data.get("timeline_date"),
            "urgency": data.get("urgency"),
            "lead_score": data.get("lead_score"),
            "notes": data.get("notes"),
            "tenant_type": data.get("tenant_type"),
            "occupant_count": data.get("occupant_count"),
            "profession": data.get("profession"),
            "workplace_text": data.get("workplace_text"),
            "assigned_user_id": assigned_user_id,
        }

        if existing_id and not update_duplicates:
            skipped += 1
            continue

        contact: Contact | None = None
        if existing_id and update_duplicates:
            result = await db.execute(select(Contact).where(Contact.id == uuid.UUID(existing_id)))
            contact = result.scalar_one_or_none()
            if contact:
                for k, v in fields.items():
                    if v is not None:
                        setattr(contact, k, v)
                updated += 1

        if not contact:
            contact = Contact(**{k: v for k, v in fields.items() if v is not None or k in ("name", "phone", "roles", "stream_type")})
            db.add(contact)
            await db.flush()
            created += 1

        if cfg.get("creates_listing") and contact:
            listing_data = _listing_fields(data, cfg)
            listing_data["contact_id"] = contact.id
            listing_data["created_by_id"] = assigned_user_id
            db.add(Listing(**{k: v for k, v in listing_data.items() if v is not None or k in ("title", "stream_type", "status")}))
            listings_created += 1

    await db.commit()
    if listings_created:
        await match_all_active(db)
    return {"created": created, "updated": updated, "skipped": skipped, "listings_created": listings_created}
