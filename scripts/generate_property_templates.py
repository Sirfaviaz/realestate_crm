#!/usr/bin/env python3
"""Generate seller/landlord Excel import templates matching PROJECT DETAILS columns."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "backend" / "templates"

CONTACT = ["name", "phone", "email", "whatsapp"]
PROJECT = [
    "builder_name",
    "project_name",
    "location",
    "project_type",
    "land_area_cent",
    "launching_time",
    "completion_time",
    "total_floors",
    "parking_details",
    "bhk",
    "sqft",
    "available_floors",
    "amenities",
    "base_price",
    "car_parking_price",
    "gst_percent",
    "down_payment_percent",
    "utility_charge",
    "total_instalment",
    "total_amount",
    "price_as_of_date",
]
CRM = ["urgency", "lead_score", "notes"]

# Human-readable labels matching the original spreadsheet
LABELS = {
    "builder_name": "BUILDER",
    "project_name": "PROJECT NAME",
    "location": "LOCATION",
    "project_type": "PROJECT TYPE",
    "land_area_cent": "LAND AREA (CENT)",
    "launching_time": "LAUNCHING TIME",
    "completion_time": "COMPLETION TIME",
    "total_floors": "TOTAL FLOORS",
    "parking_details": "PARKING DETAILS",
    "bhk": "TYPE OF APARTMENT (BHKs)",
    "sqft": "SQFT OF EACH APARTMENTS",
    "available_floors": "FLOORS OF APARTMENT (AVAILABLE)",
    "amenities": "COMMON AMENITIES",
    "base_price": "BASIC PRICE FLOOR PRICE",
    "car_parking_price": "CAR PARKING PRICE",
    "gst_percent": "GST (%)",
    "down_payment_percent": "DOWN PAYMENT (%)",
    "utility_charge": "UTILITY CHARGE",
    "total_instalment": "TOTAL INSTALMENT",
    "total_amount": "TOTAL AMOUNT",
    "price_as_of_date": "AS PER DATE",
    "monthly_rent": "MONTHLY RENT",
    "security_deposit": "SECURITY DEPOSIT",
}


def write_template(role: str, extra: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = role.capitalize()
    keys = ["name", "phone", *CONTACT[2:], *PROJECT, *extra, *CRM]
    row1 = [LABELS.get(k, k.replace("_", " ").upper()) for k in keys]
    row2 = keys
    ws.append(row1)
    ws.append(row2)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    ws.freeze_panes = "A3"
    path = OUT / f"{role}_property_template.xlsx"
    wb.save(path)
    print(f"Wrote {path}")


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    write_template("seller", [])
    write_template("landlord", ["monthly_rent", "security_deposit"])


if __name__ == "__main__":
    main()
